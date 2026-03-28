from __future__ import annotations

import json
import logging
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from .config import Settings
from .logging_utils import attach_run_file_handler, get_logger
from .minio_client import download_from_minio, ensure_bucket, upload_artifact
from .observability import write_run_trace
from .profile_cache import ensure_tabular_profile
from .schemas import ArtifactMetadata, RunCreateRequest
from .workspace_manager import WorkspaceManager, WorkspacePaths

VENDOR_ROOT = Path(__file__).resolve().parents[1] / "vendor"
if str(VENDOR_ROOT) not in sys.path:
    sys.path.insert(0, str(VENDOR_ROOT))

from agent_zero.runtime import AgentRuntime  # noqa: E402


logger = get_logger("agent_executor.run")
COMPLIANCE_SUBTASK_TEMPLATE_FILENAMES = {
    "industry_penalty": "compliance_subtask_industry_penalty.txt",
    "internal_likelihood": "compliance_subtask_internal_likelihood.txt",
    "rbi_inspection_control_risk": "compliance_subtask_rbi_inspection_control_risk.txt",
}
COMPLIANCE_QUALITY_CHECK_HEADERS = [
    "Impact Rating",
    "Likelihood Rating",
    "Control Risk Rating",
    "Fiinal Compliance Risk Rating",
    "Impact Score (Highest of Score of all Impact Categories)",
    "Likelihood Score",
    "Average Control Risk Score",
]
COMPLIANCE_ROW_KEY_CANDIDATES = [
    "Sr. No.",
    "Regulation / Act Code",
    "Paragraph Reference",
    "Action - Short Name",
]


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _artifact_priority(path: Path) -> tuple[int, str]:
    lowered = path.name.lower()
    if any(token in lowered for token in {"report", "summary", "analysis", "eda", "insight", "findings"}):
        return (0, lowered)
    if path.suffix.lower() in {".md", ".html"}:
        return (1, lowered)
    if path.suffix.lower() == ".txt":
        return (2, lowered)
    if path.suffix.lower() == ".json":
        return (3, lowered)
    if path.suffix.lower() == ".csv":
        return (4, lowered)
    return (5, lowered)


def _load_answer_from_path_reference(answer_text: str, workspace: WorkspacePaths) -> str | None:
    stripped = answer_text.strip()
    if not stripped.startswith("{"):
        return None

    try:
        payload = json.loads(stripped)
    except json.JSONDecodeError:
        return None

    referenced_path = payload.get("path")
    if not isinstance(referenced_path, str) or not referenced_path.strip():
        return None

    candidate = (workspace.root / referenced_path).resolve()
    workspace_root = workspace.root.resolve()
    try:
        candidate.relative_to(workspace_root)
    except ValueError:
        return None

    if not candidate.exists() or not candidate.is_file():
        return None

    if candidate.suffix.lower() not in {".md", ".txt", ".html", ".json", ".csv"}:
        return None

    content = candidate.read_text(encoding="utf-8", errors="replace").strip()
    return content or None


def _load_best_artifact_answer(workspace: WorkspacePaths) -> str | None:
    artifact_files: list[Path] = []
    for extension in (".md", ".html", ".txt", ".json", ".csv"):
        artifact_files.extend(
            sorted(
                [
                    path
                    for path in workspace.artifacts_dir.glob(f"*{extension}")
                    if path.is_file()
                ],
                key=_artifact_priority,
            )
        )
    for path in artifact_files:
        content = path.read_text(encoding="utf-8", errors="replace").strip()
        if content:
            return content
    return None


def _is_low_quality_answer(text: str) -> bool:
    stripped = text.strip()
    if not stripped:
        return True
    lowered = stripped.lower()
    if stripped.startswith("{") and ('"path"' in stripped or '"cmd"' in stripped or '"command"' in stripped):
        return True
    if stripped.startswith("[") and stripped.endswith("]"):
        return True
    if "import pandas" in lowered or "import json" in lowered:
        return True
    if "you can open this file" in lowered or "results saved as" in lowered:
        return True
    if "artifact:" in lowered and len(stripped) < 500:
        return True
    if "saved under `artifacts/" in lowered or "saved in `artifacts/" in lowered:
        return True
    if "the results are saved" in lowered or "the report is saved" in lowered:
        return True
    if stripped.startswith("<path>") or stripped.startswith("/workspace/"):
        return True
    if stripped.lower().startswith("generated artifact") or stripped.lower().startswith("artifacts created"):
        return True
    return False


def _summarize_selected_files(request: RunCreateRequest) -> str:
    if not request.selected_files:
        return "- none"

    top_level_counts: dict[str, int] = {}
    extension_counts: dict[str, int] = {}
    sample_paths: list[str] = []

    for item in request.selected_files:
        relative_name = item.relative_path or item.file_name
        path = Path(relative_name)
        top_level = path.parts[0] if path.parts else "."
        extension = path.suffix.lower() or f".{item.file_type.lower()}"
        top_level_counts[top_level] = top_level_counts.get(top_level, 0) + 1
        extension_counts[extension] = extension_counts.get(extension, 0) + 1
        if len(sample_paths) < 12:
            sample_paths.append(relative_name)

    top_level_summary = ", ".join(
        f"{name}: {count}" for name, count in sorted(top_level_counts.items(), key=lambda entry: (-entry[1], entry[0]))
    )
    extension_summary = ", ".join(
        f"{ext}: {count}" for ext, count in sorted(extension_counts.items(), key=lambda entry: (-entry[1], entry[0]))
    )
    samples = "\n".join(f"- {path}" for path in sample_paths)

    return (
        f"- total selected files: {len(request.selected_files)}\n"
        f"- top-level layout: {top_level_summary}\n"
        f"- file types: {extension_summary}\n"
        "- full manifest is available in `metadata/selected_files.json`\n"
        f"- sample paths:\n{samples}"
    )


def _is_compliance_assessment_run(request: RunCreateRequest) -> bool:
    if request.compliance_context is not None:
        return True
    lowered_question = (request.question or "").lower()
    return any(
        token in lowered_question
        for token in {
            "compliance risk assessment",
            "compliance assessment",
            "risk assessment",
            "inherent risk",
            "control risk",
            "final compliance risk rating",
            "compliance risk rating",
            "row-by-row compliance",
            "row by row compliance",
            "assess the first row",
            "assess the first task",
            "perform a compliance risk assessment",
            "fill this workbook",
            "fill the workbook",
            "populate this workbook",
            "populate the workbook",
            "complete this workbook",
            "complete the workbook",
            "write back to the workbook",
            "update the workbook",
        }
    )


def _is_explicit_workbook_update_requested(request: RunCreateRequest) -> bool:
    lowered_question = (request.question or "").lower()
    return any(
        token in lowered_question
        for token in {
            "fill this workbook",
            "fill the workbook",
            "populate this workbook",
            "populate the workbook",
            "complete this workbook",
            "complete the workbook",
            "write back to the workbook",
            "update the workbook",
        }
    )


def _build_compliance_context(request: RunCreateRequest) -> dict:
    context = request.compliance_context.model_dump() if request.compliance_context else {}
    if not context.get("output_artifacts"):
        context["output_artifacts"] = [
            "artifacts/compliance_assessment.json",
            "artifacts/row_scoring.csv",
            "artifacts/evidence_register.csv",
            "artifacts/assessment_summary.md",
            "artifacts/row_variance_report.csv",
            "artifacts/fill_quality_checks.md",
        ]
    if not context.get("evidence_requirements"):
        context["evidence_requirements"] = [
            "Record the source file name for every populated field.",
            "Use the exact source filename exactly as it appears in the workspace, including the real extension such as .pdf, .txt, .xlsx, .xls, or .csv.",
            "Do not paraphrase, shorten, normalize, or prettify source document names in citations or evidence registers.",
            "Where possible include page number, sheet name, row, or section reference.",
            "Distinguish 'Not Found' from 'Not Applicable' and from inferred conclusions.",
            "Maintain an evidence register alongside the assessment outputs.",
            "Tie evidence to stable row keys such as Sr. No., Paragraph Reference, or Action - Short Name.",
            "Run and save a row-variance validation artifact and an evidence-quality validation artifact when scoring multiple rows.",
        ]
    return context


def _write_compliance_subtask_templates(workspace: WorkspacePaths) -> None:
    current_date = datetime.now(timezone.utc).date().isoformat()
    templates = {
        COMPLIANCE_SUBTASK_TEMPLATE_FILENAMES["industry_penalty"]: (
            "Compliance Factor Research Task\n"
            "\n"
            "You are helping IIFL Finance evaluate one factor for a compliance risk assessment.\n"
            "\n"
            "Obligation JSON:\n"
            "{{obligation_json}}\n"
            "\n"
            "Institution context:\n"
            "- IIFL has an NBFC-ML license from RBI.\n"
            "- IIFL has a Housing Finance Company CoR from NHB.\n"
            "- IIFL has an NBFC-MFI registration from RBI.\n"
            "- IIFL also has an AMFI Registration Number and a Merchant Banker license from SEBI.\n"
            "\n"
            "Assigned factor:\n"
            "\"Whether RBI has penalized other Banks/ NBFCs/ HFCs (Industry Level) for this specific obligation only?\"\n"
            "\n"
            "Task instructions:\n"
            "- Search deeply for RBI press releases, circulars, enforcement actions, or other direct regulatory evidence tied to this exact obligation.\n"
            "- Mark the factor as `Found` only when the evidence is directly applicable and clearly shows penalties for the same obligation.\n"
            "- If the match is not direct and irrefutable, return `Not Found`.\n"
            "- Prefer exact obligation matching over broad thematic similarity.\n"
            "- Use workspace evidence first, but you may use web tools for RBI or industry-level evidence when needed.\n"
            "\n"
            "Citation format rules:\n"
            "- The document or workbook name in every citation must be the exact workspace filename exactly as stored on disk, including extension.\n"
            "- Do not paraphrase, shorten, normalize, or prettify document names.\n"
            "- If the evidence comes from mirrored extracted text, cite the exact `.txt` filename you actually read.\n"
            "- If the evidence is from a PDF or mirrored PDF text, each citation must be a single line formatted exactly as: `ExactFileName.pdf|txt | Page No | Start Line Number | End Line Number`.\n"
            "- If the evidence is from an Excel workbook, each citation must be a single line formatted exactly as: `ExactWorkbookName.xlsx|xls | Sheet Name | Cell number Top left | Cell number bottom right`.\n"
            "- If no irrefutable evidence exists, citations must be an empty array.\n"
            "\n"
            "Return exactly one JSON object in this schema and nothing else:\n"
            "{\n"
            "  \"parameter\": \"Whether RBI has penalized other Banks/ NBFCs/ HFCs (Industry Level) for this specific obligation only?\",\n"
            "  \"final_output\": \"Found/Not Found\",\n"
            "  \"reasoning\": \"<clear justification>\",\n"
            "  \"citations\": [\n"
            "    \"DocumentName.pdf | Page No | Start Line Number | End Line Number\",\n"
            "    \"WorkbookName.xlsx | Sheet Name | Cell number Top left | Cell number bottom right\"\n"
            "  ]\n"
            "}\n"
        ),
        COMPLIANCE_SUBTASK_TEMPLATE_FILENAMES["internal_likelihood"]: (
            "Compliance Factor Research Task\n"
            "\n"
            "You are helping IIFL Finance evaluate one factor for a compliance risk assessment.\n"
            "\n"
            f"Today's date is: {current_date}\n"
            "\n"
            "Obligation JSON:\n"
            "{{obligation_json}}\n"
            "\n"
            "Institution context:\n"
            "- IIFL has an NBFC-ML license from RBI.\n"
            "- IIFL has a Housing Finance Company CoR from NHB.\n"
            "- IIFL has an NBFC-MFI registration from RBI.\n"
            "- IIFL also has an AMFI Registration Number and a Merchant Banker license from SEBI.\n"
            "\n"
            "Assigned factor:\n"
            "\"Whether a non compliance event has occurred in context of this obligation?\"\n"
            "\n"
            "Task instructions:\n"
            "- Search internal evidence deeply, including compliance testing reports, audit observations, internal trackers, department clarifications, and similar IIFL evidence.\n"
            "- If a relevant non-compliance event exists and is less than 6 months old, return `High Likelihood`.\n"
            "- If a relevant non-compliance event exists and is within 6-12 months, return `Medium Likelihood`.\n"
            "- In all other cases, return `Low Likelihood`.\n"
            "- Use event dates from the evidence; do not guess dates.\n"
            "\n"
            "Citation format rules:\n"
            "- The document or workbook name in every citation must be the exact workspace filename exactly as stored on disk, including extension.\n"
            "- Do not paraphrase, shorten, normalize, or prettify document names.\n"
            "- If the evidence comes from mirrored extracted text, cite the exact `.txt` filename you actually read.\n"
            "- If the evidence is from a PDF or mirrored PDF text, each citation must be a single line formatted exactly as: `ExactFileName.pdf|txt | Page No | Start Line Number | End Line Number`.\n"
            "- If the evidence is from an Excel workbook, each citation must be a single line formatted exactly as: `ExactWorkbookName.xlsx|xls | Sheet Name | Cell number Top left | Cell number bottom right`.\n"
            "- If evidence is insufficient, citations must be an empty array.\n"
            "\n"
            "Return exactly one JSON object in this schema and nothing else:\n"
            "{\n"
            "  \"parameter\": \"Whether a non compliance event has occurred in context of this obligation?\",\n"
            "  \"final_output\": \"High Likelihood/Medium Likelihood/Low Likelihood\",\n"
            "  \"reasoning\": \"<clear justification>\",\n"
            "  \"citations\": [\n"
            "    \"DocumentName.pdf | Page No | Start Line Number | End Line Number\",\n"
            "    \"WorkbookName.xlsx | Sheet Name | Cell number Top left | Cell number bottom right\"\n"
            "  ]\n"
            "}\n"
        ),
        COMPLIANCE_SUBTASK_TEMPLATE_FILENAMES["rbi_inspection_control_risk"]: (
            "Compliance Factor Research Task\n"
            "\n"
            "You are helping IIFL Finance evaluate one factor for a compliance risk assessment.\n"
            "\n"
            f"Today's date is: {current_date}\n"
            "\n"
            "Obligation JSON:\n"
            "{{obligation_json}}\n"
            "\n"
            "Institution context:\n"
            "- IIFL has an NBFC-ML license from RBI.\n"
            "- IIFL has a Housing Finance Company CoR from NHB.\n"
            "- IIFL has an NBFC-MFI registration from RBI.\n"
            "- IIFL also has an AMFI Registration Number and a Merchant Banker license from SEBI.\n"
            "\n"
            "Assigned factor:\n"
            "\"Compliance Observations in RBI Inspection\"\n"
            "\n"
            "Task instructions:\n"
            "- Analyze RBI inspection reports and closely related internal evidence for findings related to this obligation.\n"
            "- If you find repeat or regulator-sensitive findings, return `High`.\n"
            "- If you find other relevant findings but not repeat/regulator-sensitive findings, return `Medium`.\n"
            "- If there are no relevant observations, return `Low`.\n"
            "- Focus on recurrence, regulator sensitivity, and direct relevance to the assigned obligation.\n"
            "\n"
            "Citation format rules:\n"
            "- The document or workbook name in every citation must be the exact workspace filename exactly as stored on disk, including extension.\n"
            "- Do not paraphrase, shorten, normalize, or prettify document names.\n"
            "- If the evidence comes from mirrored extracted text, cite the exact `.txt` filename you actually read.\n"
            "- If the evidence is from a PDF or mirrored PDF text, each citation must be a single line formatted exactly as: `ExactFileName.pdf|txt | Page No | Start Line Number | End Line Number`.\n"
            "- If the evidence is from an Excel workbook, each citation must be a single line formatted exactly as: `ExactWorkbookName.xlsx|xls | Sheet Name | Cell number Top left | Cell number bottom right`.\n"
            "- If evidence is insufficient, citations must be an empty array.\n"
            "\n"
            "Return exactly one JSON object in this schema and nothing else:\n"
            "{\n"
            "  \"parameter\": \"Compliance Observations in RBI Inspection\",\n"
            "  \"final_output\": \"High/Medium/Low\",\n"
            "  \"reasoning\": \"<clear justification>\",\n"
            "  \"citations\": [\n"
            "    \"DocumentName.pdf | Page No | Start Line Number | End Line Number\",\n"
            "    \"WorkbookName.xlsx | Sheet Name | Cell number Top left | Cell number bottom right\"\n"
            "  ]\n"
            "}\n"
        ),
    }

    for file_name, content in templates.items():
        (workspace.metadata_dir / file_name).write_text(content, encoding="utf-8")


def _write_compliance_templates(workspace: WorkspacePaths, request: RunCreateRequest, manifest: list[dict]) -> None:
    context = _build_compliance_context(request)
    output_artifacts = context.get("output_artifacts") or []
    evidence_requirements = context.get("evidence_requirements") or []
    scoring_instructions = context.get("scoring_instructions") or (
        "Inspect the source structure first. Derive any relevant scoring fields from the workbook or source files themselves instead of assuming a fixed schema. "
        "When scoring is required, use Found/Not Found for impact checks, derive Impact Score from the highest triggered category, "
        "derive Likelihood from recent incidents and control automation, and score rows independently from row-specific evidence."
    )

    contract = {
        "task_type": "compliance_risk_assessment",
        "selected_files": manifest,
        "output_artifacts": output_artifacts,
        "evidence_requirements": evidence_requirements,
        "scoring_instructions": scoring_instructions,
    }
    (workspace.metadata_dir / "compliance_task_contract.json").write_text(
        json.dumps(contract, indent=2),
        encoding="utf-8",
    )

    assessment_template = {
        "row_keys": {
            "primary_identifier": "",
            "secondary_identifier": "",
        },
        "target_fields": [],
        "evidence": [
            {
                "field": "",
                "value": "",
                "source_file": "",
                "source_location": "",
                "finding": "",
                "reasoning": "",
            }
        ],
        "scoring": {
            "impact_flags": {},
            "impact_score": "",
            "impact_rating": "",
            "likelihood_score": "",
            "likelihood_rating": "",
            "average_control_risk_score": "",
            "control_risk_rating": "",
            "final_compliance_risk_rating": "",
        },
    }
    (workspace.analysis_dir / "compliance_row_template.json").write_text(
        json.dumps(assessment_template, indent=2),
        encoding="utf-8",
    )

    instructions = [
        "Compliance Assessment Contract",
        "",
        "Target sources:",
        "- inspect the selected files and determine the source workbook(s), PDFs, trackers, and other materials relevant to the requested assessment scope",
        "Expected output artifacts:",
        *[f"- {artifact}" for artifact in output_artifacts],
        "",
        "Source handling:",
        "- treat spreadsheets such as tasks.xlsx as structured source material for row-level assessment",
        "- do not assume a fixed header list from the prompt",
        "- inspect the source structure and derive the actual fields, row keys, and evidence locations from the files themselves",
        "- do not turn the task into workbook-filling work unless the user explicitly asks for workbook updates",
        "",
        "Evidence requirements:",
        *[f"- {item}" for item in evidence_requirements],
        "",
        "Scoring instructions:",
        f"- {scoring_instructions}",
        "- Impact Score must be the highest triggered score across the five impact categories.",
        "- Impact Rating should follow the numeric score.",
        "- Likelihood must use event recency and control design/automation evidence.",
        "- Average Control Risk Score must be numeric and derived from the factor ratings.",
        "- Final assessment outputs should contain concise values, while detailed reasoning should go to the evidence register artifact and assessment summary.",
        "- Score each assessed row independently using that row's obligation text and evidence; do not propagate one global profile across rows.",
        "- Do not hardcode one shared Impact/Likelihood/Control Risk/Final Risk value for the entire assessment unless the source material explicitly proves that every row truly shares the same result.",
        "- Do not hardcode dataset-level constants, blanket defaults, or script-level fallback values that are then reused across most rows.",
        "- Each evidence-register entry must be tied to a specific assessed row using stable row keys.",
        "- If `metadata/compliance_subtask_*.txt` templates are present, you must run the corresponding Task-tool factor subtasks before writing broad assessment scripts or finalizing assessment outputs.",
        "- When those compliance subtask templates are present, do not skip delegation and do not satisfy the assessment solely by writing one monolithic main-agent script.",
        "- Before finalizing a multi-row assessment, run a self-check for suspiciously uniform outputs across the core risk-rating columns.",
        "- Save the self-check outputs as `artifacts/row_variance_report.csv` and `artifacts/fill_quality_checks.md`.",
        "- `artifacts/fill_quality_checks.md` must explicitly state whether the assessment passed row-variance review and evidence-quality review.",
    ]
    (workspace.analysis_dir / "compliance_assessment_instructions.md").write_text(
        "\n".join(instructions),
        encoding="utf-8",
    )

    validation_requirements = [
        "Compliance Validation Requirements",
        "",
        "Validation is part of the deliverable. You must validate your own assessment before finishing.",
        "",
        "Required validation outputs:",
        "- artifacts/row_variance_report.csv",
        "- artifacts/fill_quality_checks.md",
        "",
        "Validation rules:",
        "- Report non-empty row count, unique value count, dominant value share, and dominant value for each core scoring column.",
        "- Core scoring columns: Impact Score (Highest of Score of all Impact Categories), Impact Rating, Likelihood Score, Likelihood Rating, Average Control Risk Score, Control Risk Rating, Fiinal Compliance Risk Rating.",
        "- Flag any column that is suspiciously uniform across most rows.",
        "- Review the evidence register for repeated evidence text reused across many different rows and call that out explicitly.",
        "- Review citations and evidence-register source_file values for exact on-disk filenames including extensions; flag paraphrased or normalized document names.",
        "- If outputs are suspiciously uniform, revise the assessment instead of merely noting the issue.",
        "",
        "Hard prohibitions:",
        "- Do not use one global risk profile for all rows.",
        "- Do not use hardcoded constants as substitutes for row-level analysis.",
        "- Do not claim validation passed unless you actually generated and inspected the validation artifacts.",
    ]
    (workspace.analysis_dir / "compliance_validation_requirements.md").write_text(
        "\n".join(validation_requirements),
        encoding="utf-8",
    )
    _write_compliance_subtask_templates(workspace)


def _build_prompt(request: RunCreateRequest, workspace: WorkspacePaths) -> str:
    files_manifest = _summarize_selected_files(request)
    history = "\n".join(f"{message.role.upper()}: {message.content}" for message in request.history[-6:])
    prior_context = request.prior_context or "- none"
    compliance_mode = _is_compliance_assessment_run(request)
    compliance_context = _build_compliance_context(request) if compliance_mode else {}
    compliance_block = ""
    if compliance_mode:
        evidence_lines = "\n".join(f"- {item}" for item in compliance_context.get("evidence_requirements", []))
        artifacts = "\n".join(f"- {item}" for item in compliance_context.get("output_artifacts", []))
        compliance_block = (
            "Compliance assessment mode is active.\n"
            "- Treat workbooks such as tasks.xlsx as structured source material for row-level compliance assessment.\n"
            "- Inspect the source structure first and derive the relevant row keys, scoring fields, and evidence locations from the source files themselves.\n"
            "- Treat each workbook row as an independent obligation unless the source material explicitly indicates otherwise.\n"
            "- Do not assume the short file summary in this prompt is exhaustive; inspect `metadata/selected_files.json` and the workspace tree for the full file list.\n"
            "- Produce structured assessment outputs, not just a narrative answer.\n"
            "- Use `analysis/compliance_assessment_instructions.md` and `analysis/compliance_row_template.json` as the contract.\n"
            "- Use `analysis/compliance_validation_requirements.md` to validate your own assessment before finishing.\n"
            "- Reusable delegated-task templates are available in `metadata/compliance_subtask_industry_penalty.txt`, `metadata/compliance_subtask_internal_likelihood.txt`, and `metadata/compliance_subtask_rbi_inspection_control_risk.txt`.\n"
            "- When those compliance subtask templates are present, you must use the Task tool with subagent `compliance_factor_analyst` for the available factor-level research before broad synthesis.\n"
            "- For delegated factor analysis, replace `{{obligation_json}}` in the template with the row-specific obligation JSON and ask for exactly one JSON object back.\n"
            "- Launch the available factor-analysis subtasks early, preferably in parallel, after you have identified the row-specific obligation JSON.\n"
            "- Do not skip the required Task-tool delegation just because you can write a custom analysis script in the main agent.\n"
            "- Do not satisfy the compliance assessment only by writing one monolithic script when the delegated-task templates are present.\n"
            "- You remain responsible for the final assessment outputs, evidence register, and final synthesis.\n"
            "- Save the final assessment artifacts as:\n"
            f"{artifacts}\n"
            "- Also create an evidence register showing the source file and source location for every populated field.\n"
            "- Every citation and every evidence-register source file value must use the exact workspace filename exactly as stored on disk, including extension.\n"
            "- Do not paraphrase, shorten, normalize, or prettify document names; if you read a mirrored extracted-text file, cite that exact `.txt` filename.\n"
            "- For inherent impact, explicitly evaluate these five checks and store Found/Not Found values with evidence:\n"
            "  1. RBI penalties on other Banks/NBFCs/HFCs for the specific obligation.\n"
            "  2. RBI embargo on business for the specific obligation.\n"
            "  3. RBI cancellation of license for the specific obligation.\n"
            "  4. Whether the guideline has customer impact.\n"
            "  5. Whether there are additional RBI requirements not covered above.\n"
            "- If any of checks 1-4 are Found, Impact is High. If only check 5 is Found, Impact is Medium. Otherwise Impact is Low.\n"
            "- For likelihood, use recency of non-compliance events plus control existence and automation level.\n"
            "- For control risk, fill every control-risk field, map ratings to numbers consistently, calculate the average score, then derive Control Risk Rating and Final Compliance Risk Rating.\n"
            "- Do not use one global industry summary, one global internal summary, or one constant script default to populate all rows.\n"
            "- Do not hardcode dataset-level counts, static score defaults, or one-size-fits-all fallback values and then reuse them across most rows.\n"
            "- If many rows end up with the same rating, that is acceptable only if you can justify it row-by-row and your evidence register still contains row-specific evidence.\n"
            "- For multi-row assessments, your quality bar is not merely that outputs are non-empty; the assessment must withstand a row-variance sanity check on the core scoring columns and an evidence-quality check for repeated boilerplate.\n"
            "- Before finalizing a multi-row assessment, generate `artifacts/row_variance_report.csv` and `artifacts/fill_quality_checks.md`, inspect them, and revise the assessment if they indicate suspicious uniformity.\n"
            "- Where an output field needs a compact value, keep the detailed reasoning in the evidence register and scoring summary.\n"
            "- Do not stop at a markdown report if the requested assessment outputs can be produced.\n"
            "- Evidence requirements:\n"
            f"{evidence_lines}\n\n"
        )
    return (
        "You are analyzing files in an isolated workspace as a data analyst.\n\n"
        f"Workspace directories:\n"
        f"- input/: source files (spreadsheets, CSVs, PDFs)\n"
        f"- profiles/: optional deterministic schema/profile artifacts when available\n"
        f"- cache/: optional reusable cleaned tables, extracted PDF text, and prior artifacts from earlier turns\n"
        f"- analysis/: scripts and intermediate outputs\n"
        f"- artifacts/: final generated outputs\n\n"
        f"Selected files summary:\n{files_manifest or '- none'}\n\n"
        f"Recent conversation:\n{history or '- none'}\n\n"
        f"Prior thread context:\n{prior_context}\n\n"
        f"User question:\n{request.question}\n\n"
        "CRITICAL: Treat the workspace files as the primary source of truth.\n"
        f"{compliance_block}"
        "- NEVER run `pip install`, `pip3 install`, `conda install`, or any package manager command.\n"
        "- All required packages are pre-installed. Use only: pandas, numpy, openpyxl, xlsxwriter, xlrd, pyarrow, "
        "matplotlib, seaborn, plotly, kaleido, scipy, statsmodels, scikit-learn, "
        "pdfplumber, pymupdf (fitz), pdfminer.six, pikepdf, "
        "nltk, regex, beautifulsoup4, lxml, Pillow, Jinja2, python-pptx, python-docx, "
        "markdown, pyyaml, jsonschema, rich, tqdm, sympy, tabulate, chardet.\n\n"
        "Operating requirements:\n"
        "- inspect the selected input files before answering\n"
        "- if profiles/*.json or cache/cleaned/ already exist, reuse them; otherwise work directly from raw files in input/\n"
        "- if `input/extracted_text_from_pdf/` or `extracted_text_from_pdf/` exists, prefer those mirrored `.txt` extractions before re-reading raw PDFs\n"
        "- for PDFs, use pdfplumber or pymupdf (import fitz) for extraction when needed\n"
        "- for follow-up questions in the same thread, reuse existing profiles, cleaned tables, and prior artifacts in cache/ whenever they already answer part of the task\n"
        "- before writing any inspection script for spreadsheets, check profiles/*.json first when present\n"
        "- for spreadsheets, inspect workbook structure first: sheet names, header rows, columns, row counts, missing values, date/time fields, numeric fields, units, currencies, and percentage columns\n"
        "- if a user refers to a specific sheet, annexure, tab, company, period, row range, metric, category, or PDF page/section, inspect that exact content before answering\n"
        "- if the user provides a URL or explicitly asks for internet research, you may use webfetch or websearch; fetch the exact URL first when provided\n"
        "- when using internet sources, clearly distinguish web-derived findings from workspace-derived findings\n"
        "- when analysis is needed, write reproducible scripts under analysis/ and run them\n"
        "- avoid repeatedly rewriting similar scripts; prefer adapting cached cleaned tables or writing one focused script per task\n"
        "- do not modify files in input/\n"
        "- do not modify files in profiles/ or cache/cleaned unless explicitly required; treat them as reusable inputs\n"
        "- save user-facing outputs under artifacts/\n"
        "- if the user asks for graphs, charts, plots, dashboards, or visualizations, always save them under artifacts/\n"
        "- if the user asks for a report, save a report artifact under artifacts/ in an appropriate format such as markdown, html, json, csv, or presentation output when suitable\n"
        "- when compliance assessment mode is active, treat structured assessment outputs and evidence capture as mandatory deliverables\n"
        "- if you perform EDA or statistics, include schema/profile findings, missing-data notes, sample size, assumptions, and caveats where relevant\n"
        "- if you perform correlation, regression, or similar statistics, clearly identify the variables used and how to interpret the result\n"
        "- if you compare multiple workbooks or sheets, explicitly describe the mapping and alignment used\n"
        "- after generating an artifact, read the relevant result back and answer the user's question directly in natural language\n"
        "- mention generated artifact filenames in the final answer when you create them\n"
        "- do not use emojis in the final answer\n"
        "- do not embed raw workspace-relative image paths like artifacts/foo.png in markdown image tags or html because the chat UI will not render them inline\n"
        "- do not end with only a file path, a JSON payload, raw code, or a note that results were saved\n"
    )


def _stage_selected_files(
    request: RunCreateRequest,
    settings: Settings,
    workspace: WorkspacePaths,
    workspace_manager: WorkspaceManager,
    emit: Callable[[str, dict], None],
    run_logger: logging.LoggerAdapter,
) -> list[dict]:
    def _replace_with_copy(source: Path, destination: Path) -> Path:
        destination.parent.mkdir(parents=True, exist_ok=True)
        if destination.exists() or destination.is_symlink():
            if destination.is_dir() and not destination.is_symlink():
                shutil.rmtree(destination)
            else:
                destination.unlink()
        shutil.copytree(source, destination)
        return destination

    manifest: list[dict] = []
    for selected_file in request.selected_files:
        relative_input_path = Path(selected_file.relative_path or selected_file.file_name)
        shared_path = workspace.shared_input_dir / relative_input_path
        local_destination = workspace.input_dir / relative_input_path

        if selected_file.workspace_staged_path:
            staged_source = settings.workspace_root / selected_file.workspace_staged_path
            if staged_source.exists():
                if staged_source.is_dir():
                    _replace_with_copy(staged_source, shared_path)
                    local_path = _replace_with_copy(shared_path, local_destination)
                    run_logger.info(
                        "copied staged workspace directory file_name=%s staged_source=%s relative_input_path=%s",
                        selected_file.file_name,
                        staged_source,
                        relative_input_path,
                    )
                else:
                    local_path = workspace_manager.sync_into_run(staged_source, local_destination)
                    shared_path.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(staged_source, shared_path)
                    run_logger.info(
                        "staged imported file file_name=%s staged_source=%s relative_input_path=%s",
                        selected_file.file_name,
                        staged_source,
                        relative_input_path,
                    )
                emit(
                    "file.reused",
                    {
                        "run_id": request.run_id,
                        "file_id": selected_file.file_id,
                        "file_name": selected_file.file_name,
                        "local_path": str(local_path.relative_to(workspace.root)),
                    },
                )
            else:
                raise FileNotFoundError(f"Staged workspace file not found: {staged_source}")
        elif shared_path.exists():
            local_path = workspace_manager.sync_into_run(shared_path, local_destination)
            run_logger.info("reused cached file file_name=%s shared_path=%s", selected_file.file_name, shared_path)
            emit(
                "file.reused",
                {
                    "run_id": request.run_id,
                    "file_id": selected_file.file_id,
                    "file_name": selected_file.file_name,
                    "local_path": str(local_path.relative_to(workspace.root)),
                },
            )
        else:
            run_logger.info("downloading file file_id=%s file_name=%s", selected_file.file_id, selected_file.file_name)
            download_from_minio(selected_file.file_url, shared_path)
            local_path = workspace_manager.sync_into_run(shared_path, local_destination)
            run_logger.info("downloaded file file_name=%s local_path=%s", selected_file.file_name, local_path)
            emit(
                "file.downloaded",
                {
                    "run_id": request.run_id,
                    "file_id": selected_file.file_id,
                    "file_name": selected_file.file_name,
                    "local_path": str(local_path.relative_to(workspace.root)),
                },
            )

        manifest.append(
            {
                "file_id": selected_file.file_id,
                "file_name": selected_file.file_name,
                "relative_path": str(relative_input_path),
                "local_path": str(local_path),
            }
        )

        if not selected_file.preprocessing_skipped:
            profile_path = ensure_tabular_profile(shared_path, selected_file.file_type, workspace.shared_profiles_dir, workspace.shared_cache_dir)
            workspace_manager.sync_into_run(profile_path, workspace.profiles_dir / profile_path.name)

            cleaned_root = workspace.shared_cache_dir / "cleaned" / selected_file.file_name
            if cleaned_root.exists():
                destination_root = workspace.cache_dir / "cleaned" / selected_file.file_name
                if destination_root.exists():
                    shutil.rmtree(destination_root)
                shutil.copytree(cleaned_root, destination_root)

            emit(
                "profile.ready",
                {
                    "run_id": request.run_id,
                    "file_id": selected_file.file_id,
                    "file_name": selected_file.file_name,
                    "profile_path": str((workspace.profiles_dir / profile_path.name).relative_to(workspace.root)),
                },
            )
    return manifest


def _stage_prior_artifacts(
    request: RunCreateRequest,
    workspace: WorkspacePaths,
    emit: Callable[[str, dict], None],
    run_logger: logging.LoggerAdapter,
) -> list[dict]:
    staged: list[dict] = []
    for artifact in request.prior_artifacts:
        relative_path = artifact.workspace_path or artifact.name
        destination = workspace.prior_artifacts_dir / relative_path
        if destination.exists():
            staged.append({"name": artifact.name, "workspace_path": str(destination.relative_to(workspace.root)), "artifact_type": artifact.artifact_type})
            continue

        download_from_minio(artifact.url, destination)
        staged_item = {
            "name": artifact.name,
            "workspace_path": str(destination.relative_to(workspace.root)),
            "artifact_type": artifact.artifact_type,
        }
        staged.append(staged_item)
        run_logger.info("staged prior artifact name=%s workspace_path=%s", artifact.name, staged_item["workspace_path"])
        emit(
            "artifact.reused",
            {
                "run_id": request.run_id,
                "artifact": staged_item,
            },
        )
    return staged


def _find_header_row(worksheet) -> tuple[int, dict[str, int]]:
    required = set(COMPLIANCE_QUALITY_CHECK_HEADERS)
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        header_map: dict[str, int] = {}
        for column_index, value in enumerate(row):
            if value is None:
                continue
            header = str(value).strip()
            if header:
                header_map[header] = column_index
        if required.intersection(header_map):
            return row_index, header_map
    return 0, {}


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.6f}".rstrip("0").rstrip(".")
    return str(value).strip()


def _write_quality_report(path: Path, lines: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _validate_compliance_workbook_outputs(workspace: WorkspacePaths, request: RunCreateRequest, run_logger: logging.LoggerAdapter) -> None:
    workbook_candidates = []
    workbook_candidates.extend(sorted(workspace.artifacts_dir.glob("*filled*.xlsx")))
    workbook_candidates.extend(sorted(workspace.artifacts_dir.glob("*workbook*.xlsx")))
    workbook_candidates.extend(sorted(workspace.artifacts_dir.glob("*.xlsx")))

    workbook_path = next((path for path in workbook_candidates if path.exists() and path.is_file()), None)
    if workbook_path is None:
        run_logger.warning("compliance quality check skipped: no workbook artifact found")
        return

    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row_index, header_map = _find_header_row(worksheet)
    if not header_map:
        run_logger.warning("compliance quality check skipped: could not identify workbook headers in %s", workbook_path)
        return

    key_indexes = [header_map[key] for key in COMPLIANCE_ROW_KEY_CANDIDATES if key in header_map]
    data_rows: list[dict[str, str]] = []
    for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        row_values = [_normalize_cell(value) for value in row]
        if not any(row_values):
            continue
        if key_indexes and not any(row_values[index] for index in key_indexes if index < len(row_values)):
            continue
        data_rows.append(
            {
                header: row_values[index] if index < len(row_values) else ""
                for header, index in header_map.items()
            }
        )

    report_lines = [
        "# Compliance Fill Quality Checks",
        "",
        f"- Workbook: `{workbook_path.name}`",
        f"- Worksheet: `{worksheet.title}`",
        f"- Header row index: {header_row_index}",
        f"- Data rows inspected: {len(data_rows)}",
        "",
        "## Core Column Variance",
        "",
        "| Column | Non-empty rows | Unique values | Dominant value share | Dominant value |",
        "| --- | ---: | ---: | ---: | --- |",
    ]

    suspicious_columns: list[str] = []
    for header in COMPLIANCE_QUALITY_CHECK_HEADERS:
        values = [row.get(header, "") for row in data_rows if row.get(header, "")]
        if not values:
            report_lines.append(f"| {header} | 0 | 0 | n/a |  |")
            continue
        counts: dict[str, int] = {}
        for value in values:
            counts[value] = counts.get(value, 0) + 1
        dominant_value, dominant_count = max(counts.items(), key=lambda item: item[1])
        dominant_share = dominant_count / len(values)
        report_lines.append(
            f"| {header} | {len(values)} | {len(counts)} | {dominant_share:.1%} | {dominant_value} |"
        )
        if len(values) >= 10 and len(counts) == 1 and dominant_share >= 0.95:
            suspicious_columns.append(header)

    report_path = workspace.artifacts_dir / "fill_quality_checks.md"
    if suspicious_columns:
        report_lines.extend(
            [
                "",
                "## Result",
                "",
                "FAILED",
                "",
                "The workbook looks suspiciously uniform across core scoring columns. This usually means the agent reused one global profile instead of scoring rows independently.",
                "",
                "Suspicious columns:",
                *[f"- {header}" for header in suspicious_columns],
            ]
        )
        _write_quality_report(report_path, report_lines)
        raise ValueError(
            "Compliance workbook quality check failed: suspiciously uniform values across core risk columns "
            f"({', '.join(suspicious_columns)}). See artifacts/fill_quality_checks.md."
        )

    report_lines.extend(
        [
            "",
            "## Result",
            "",
            "PASSED",
            "",
            "No suspiciously uniform pattern was detected across the core scoring columns.",
        ]
    )
    _write_quality_report(report_path, report_lines)
    run_logger.info("compliance quality check passed workbook=%s rows=%s", workbook_path.name, len(data_rows))


def _map_runtime_event(event: dict, emit: Callable[[str, dict], None], run_id: str) -> None:
    event_type = event.get("type")
    if event_type == "step.start":
        emit("agent.step", {"run_id": run_id, "step": event.get("step"), "message_count": event.get("message_count")})
    elif event_type == "tool.start":
        emit(
            "agent.tool_call",
            {
                "run_id": run_id,
                "step": event.get("step"),
                "tool": event.get("name"),
                "args": event.get("args", {}),
                "source": event.get("source"),
            },
        )
    elif event_type == "tool.end":
        emit(
            "agent.tool_result",
            {
                "run_id": run_id,
                "step": event.get("step"),
                "tool": event.get("name"),
                "title": event.get("title"),
                "output_preview": event.get("output_preview"),
                "metadata": event.get("metadata", {}),
            },
        )
    elif event_type == "tool.error":
        emit(
            "agent.tool_result",
            {
                "run_id": run_id,
                "step": event.get("step"),
                "tool": event.get("name"),
                "error": event.get("error"),
            },
        )
    elif event_type == "run.finish":
        emit(
            "agent.summary",
            {
                "run_id": run_id,
                "reason": event.get("reason"),
                "step": event.get("step"),
                "tool_events": event.get("tool_events"),
                "text_preview": event.get("text_preview"),
            },
        )


def execute_run(
    request: RunCreateRequest,
    settings: Settings,
    workspace_manager: WorkspaceManager,
    emit: Callable[[str, dict], None],
) -> tuple[str, list[ArtifactMetadata], list[dict], str, str | None]:
    workspace = workspace_manager.create(request.thread_id, request.run_id)
    settings.trace_root.mkdir(parents=True, exist_ok=True)
    log_path = workspace.logs_dir / "run.log"
    file_handler = attach_run_file_handler(logger, log_path)
    run_logger = logging.LoggerAdapter(
        logger,
        {"run_id": request.run_id, "thread_id": request.thread_id},
    )
    run_logger.info(
        "run started question=%r selected_files=%s workspace=%s",
        request.question,
        [item.file_name for item in request.selected_files],
        workspace.root,
    )
    emit("workspace.created", {"run_id": request.run_id, "path": str(workspace.root)})
    started_at = _utc_now()
    execution_trace: list[dict] = []
    uploaded: list[ArtifactMetadata] = []
    answer_text = ""
    try:
        manifest = _stage_selected_files(request, settings, workspace, workspace_manager, emit, run_logger)
        staged_prior_artifacts = _stage_prior_artifacts(request, workspace, emit, run_logger)
        (workspace.metadata_dir / "selected_files.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
        (workspace.metadata_dir / "prior_artifacts.json").write_text(json.dumps(staged_prior_artifacts, indent=2), encoding="utf-8")
        (workspace.metadata_dir / "prior_context.txt").write_text(request.prior_context or "", encoding="utf-8")
        if _is_compliance_assessment_run(request):
            _write_compliance_templates(workspace, request, manifest)
        prompt = _build_prompt(request, workspace)
        (workspace.metadata_dir / "prompt.txt").write_text(prompt, encoding="utf-8")
        run_logger.info(
            "prepared prompt history_items=%s prior_artifacts=%s profiles=%s",
            len(request.history),
            len(staged_prior_artifacts),
            len(list(workspace.profiles_dir.glob("*.json"))),
        )

        runtime = AgentRuntime(
            repo_root=VENDOR_ROOT / "agent_zero",
            workspace_root=workspace.root,
            cwd=workspace.root,
            interactive=False,
            api_base=settings.openai_base_url,
            request_timeout=settings.request_timeout,
        )

        def on_event(event: dict) -> None:
            execution_trace.append(event)
            event_type = event.get("type", "unknown")
            if event_type in {"step.start", "tool.start", "tool.end", "tool.error", "run.finish"}:
                run_logger.info("runtime_event type=%s payload=%s", event_type, event)
            _map_runtime_event(event, emit, request.run_id)

        run_logger.info(
            "invoking agent runtime agent=%s model=%s max_steps=%s api_base=%s",
            "workspace_analyst",
            request.model or settings.model_name,
            settings.max_steps,
            settings.openai_base_url,
        )
        result = runtime.run(
            user_prompt=prompt,
            agent="workspace_analyst",
            model=request.model or settings.model_name,
            max_steps=settings.max_steps,
            on_event=on_event,
        )
        answer_text = result.text
        run_logger.info("agent runtime completed answer_chars=%s", len(answer_text))

        if _is_explicit_workbook_update_requested(request):
            _validate_compliance_workbook_outputs(workspace, request, run_logger)

        ensure_bucket(settings.artifacts_bucket)
        discovered = workspace_manager.discover_artifacts(workspace)
        run_logger.info("discovered artifacts count=%s", len(discovered))
        for artifact in discovered:
            local_path = Path(artifact.local_path or "")
            workspace_path = artifact.workspace_path or artifact.name
            object_key = f"workspace-agent/{request.thread_id}/{request.run_id}/{workspace_path}"
            artifact_url = upload_artifact(local_path, settings.artifacts_bucket, object_key)
            uploaded.append(
                ArtifactMetadata(
                    name=artifact.name,
                    workspace_path=artifact.workspace_path,
                    artifact_type=artifact.artifact_type,
                    url=artifact_url,
                    mime_type=artifact.mime_type,
                    size_bytes=artifact.size_bytes,
                    local_path=artifact.local_path,
                )
            )
            run_logger.info("uploaded artifact name=%s object_key=%s", artifact.name, object_key)
            emit(
                "artifact.created",
                {
                    "run_id": request.run_id,
                    "artifact": uploaded[-1].model_dump(),
                },
            )

        referenced_answer = _load_answer_from_path_reference(answer_text, workspace)
        if referenced_answer:
            run_logger.warning("replacing path-reference final answer with file content")
            answer_text = referenced_answer

        if _is_low_quality_answer(answer_text):
            fallback_answer = _load_best_artifact_answer(workspace)
            if fallback_answer:
                run_logger.warning("replacing low-quality final answer with artifact content")
                answer_text = fallback_answer

        (workspace.logs_dir / "execution_trace.json").write_text(json.dumps(execution_trace, indent=2, default=str), encoding="utf-8")
        (workspace.logs_dir / "final_answer.md").write_text(answer_text, encoding="utf-8")
        trace_dir, trace_urls = write_run_trace(
            trace_root=settings.trace_root,
            trace_bucket=settings.trace_bucket,
            request=request,
            settings={
                "model_name": settings.model_name,
                "openai_base_url": settings.openai_base_url,
                "request_timeout": settings.request_timeout,
                "max_steps": settings.max_steps,
            },
            execution_trace=execution_trace,
            uploaded=uploaded,
            answer_text=answer_text,
            log_path=str(log_path),
            status="completed",
            started_at=started_at,
            finished_at=_utc_now(),
        )
        run_logger.info("run completed successfully log_path=%s trace_dir=%s", log_path, trace_dir)
        return answer_text, uploaded, execution_trace, str(log_path), trace_urls.get("summary.json")
    except Exception as exc:
        trace_dir, trace_urls = write_run_trace(
            trace_root=settings.trace_root,
            trace_bucket=settings.trace_bucket,
            request=request,
            settings={
                "model_name": settings.model_name,
                "openai_base_url": settings.openai_base_url,
                "request_timeout": settings.request_timeout,
                "max_steps": settings.max_steps,
            },
            execution_trace=execution_trace,
            uploaded=uploaded,
            answer_text=answer_text,
            log_path=str(log_path),
            status="failed",
            started_at=started_at,
            finished_at=_utc_now(),
            error=str(exc),
        )
        run_logger.exception("run failed trace_dir=%s", trace_dir)
        raise
    finally:
        logger.removeHandler(file_handler)
        file_handler.close()
